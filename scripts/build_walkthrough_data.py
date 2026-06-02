#!/usr/bin/env python3
"""Dump the client's master workbook → src/data/walkthroughWorkbook.json.

The Client Walkthrough renders the client file 1:1, so we extract all
three data sheets:
  - INVENTORY (rollup, with the audit colour classifications)
  - IMEI NUMBERS (per-unit register)
  - SUPPLIER WHATSAPP UPDATES (free-text feed)

The INVENTORY sheet keeps its per-row classification (green/yellow/red
/blue) derived from the cell fills we applied during the audit pass.
The other two sheets ship as raw rows.
"""
import json
import os
from openpyxl import load_workbook
from collections import Counter

SRC = 'INVENTORY_REPORT_2026_annotated.xlsx'
OUT = 'src/data/walkthroughWorkbook.json'

FILL_CLASS = {
    '00C6EFCE': 'green',
    '00FFC7CE': 'red',
    '00FFEB9C': 'yellow',
    '00BDD7EE': 'blue',
}
DIVIDER_KEYWORDS = {'TOTAL', 'SHS', 'SOLD', 'SOLD ITEMS'}

def row_class(ws, row_i, n_cols):
    seen = Counter()
    for col in range(1, n_cols + 1):
        c = ws.cell(row_i, col)
        if not (c.fill and c.fill.patternType == 'solid'):
            continue
        rgb = c.fill.fgColor.rgb
        if rgb in FILL_CLASS:
            seen[FILL_CLASS[rgb]] += 1
    for cls in ('red', 'yellow', 'green'):
        if seen[cls] >= 3:
            return cls
    if seen['blue'] > 0:
        return 'blue'
    return 'none'

def dump_sheet(wb, sheet_name, classify=False):
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    # Drop trailing all-None header columns.
    while headers and headers[-1] is None:
        headers.pop()
    n_cols = len(headers)

    out_rows = []
    for r in range(2, ws.max_row + 1):
        cells = [ws.cell(r, c).value for c in range(1, n_cols + 1)]
        all_blank = all(v in (None, '') for v in cells)
        if all_blank:
            out_rows.append({'row': r, 'cells': cells, 'classification': 'blank'})
            continue
        klass = 'none'
        if classify:
            first = cells[0]
            if first and str(first).strip().upper() in DIVIDER_KEYWORDS:
                klass = 'divider'
            else:
                klass = row_class(ws, r, n_cols)
        out_rows.append({'row': r, 'cells': cells, 'classification': klass})

    # Strip trailing blank rows.
    while out_rows and out_rows[-1]['classification'] == 'blank':
        out_rows.pop()
    return {'headers': headers, 'rows': out_rows}

def main():
    wb = load_workbook(SRC)
    sheets = {}
    sheets['INVENTORY']                  = dump_sheet(wb, 'INVENTORY', classify=True)
    sheets['IMEI NUMBERS']               = dump_sheet(wb, 'IMEI NUMBERS')
    sheets['SUPPLIER WHATSAPP UPDATES']  = dump_sheet(wb, 'SUPPLIER WHATSAPP UPDATES')

    for name, s in sheets.items():
        counts = Counter(r['classification'] for r in s['rows'])
        print(f"{name}: {len(s['rows'])} rows  cols={len(s['headers'])}  classes={dict(counts)}")

    os.makedirs('src/data', exist_ok=True)
    with open(OUT, 'w') as f:
        json.dump(sheets, f, indent=1, default=str)
    print(f"→ saved {OUT}")

if __name__ == '__main__':
    main()
